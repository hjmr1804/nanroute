"""
app.py — Planificador de reparto: entregas + recogidas desde una base central,
con hora de inicio, ETAs, prioridades y equilibrio por tiempo.
Ejecutar:  streamlit run app.py
"""
from __future__ import annotations
import os
import pandas as pd
import numpy as np
import streamlit as st
import folium
from folium.plugins import AntPath
import streamlit.components.v1 as components

import route_core as rc
import ors
from estilos_ui import aplicar_estilos

st.set_page_config(page_title="NanRoute",
                   page_icon="favicon.png" if os.path.exists("favicon.png") else "🚚",
                   layout="wide")
aplicar_estilos()
st.markdown('<div class="credito">by: Hernando Mejía Gómez</div>', unsafe_allow_html=True)
CACHE = "geocode_cache.csv"
DEFAULT_ORS_KEY = ("eyJvcmciOiI1YjNjZTM1OTc4NTExMTAwMDFjZjYyNDgiLCJpZCI6IjNhNWQyNjFiZmZh"
                   "NjQwYmU5YTQ2NWUwZDYyZjhlNmY0IiwiaCI6Im11cm11cjY0In0=")
COLS = ["id","tipo","direccion","codigo_postal","ciudad","pais",
        "prioritario","fecha_hora_limite","lat","lon","notas"]


# ---------------- geocodificación ----------------
def load_cache():
    if os.path.exists(CACHE):
        c = pd.read_csv(CACHE)
        return {r["direccion"]: (r["lat"], r["lon"]) for _, r in c.iterrows() if pd.notna(r["lat"])}
    return {}

def save_cache(cache):
    pd.DataFrame([{"direccion":k,"lat":v[0],"lon":v[1]} for k,v in cache.items()]).to_csv(CACHE, index=False)

def _secret(name):
    try: return st.secrets.get(name, "")
    except Exception: return ""

def _norm(s):
    import unicodedata
    s = str(s or "")
    s = "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")
    return s.lower().strip()

def _val(x):
    s = str(x).strip() if x is not None else ""
    return "" if s.lower() in ("nan", "none", "") else s

def sig(row):
    cp = str(row.get("codigo_postal","")).split(".")[0].strip()
    return " | ".join([str(row.get("direccion","")).strip(), cp,
                       str(row.get("ciudad","")).strip(), str(row.get("pais","")).strip()])

def _town(loc):
    a = loc.raw.get("address", {}) if getattr(loc, "raw", None) else {}
    for k in ("city","town","village","municipality","suburb","county"):
        if a.get(k): return a[k]
    return ""

def _geoloc(email):
    from geopy.geocoders import Nominatim
    return Nominatim(user_agent=f"tesis-reparto/{email or 'demo'}", timeout=10)

def _rl(gl):
    from geopy.extra.rate_limiter import RateLimiter
    return RateLimiter(gl.geocode, min_delay_seconds=1.0, max_retries=2, error_wait_seconds=2.0)

def _city_viewbox(gcode, ciudad, pais):
    """Bounding box de la ciudad, para acotar la búsqueda y no saltar de pueblo."""
    if not str(ciudad or "").strip(): return None
    key = f"{_norm(ciudad)}|{_norm(pais)}"
    cache = st.session_state.setdefault("city_bbox", {})
    if key in cache: return cache[key]
    bb = None
    try:
        loc = gcode(f"{ciudad}, {pais or 'España'}", country_codes="es")
        if loc and loc.raw.get("boundingbox"):
            s, n, w, e = map(float, loc.raw["boundingbox"])
            bb = [(s, w), (n, e)]
    except Exception:
        bb = None
    cache[key] = bb
    return bb

def geocode_one(gcode, row, viewbox):
    """Búsqueda ESTRUCTURADA país > ciudad > calle, acotada a la ciudad."""
    q = {"country": _val(row.get("pais")) or "España"}
    street = _val(row.get("direccion"))
    if street: q["street"] = street
    ciudad = _val(row.get("ciudad"))
    if ciudad: q["city"] = ciudad
    cp = _val(row.get("codigo_postal")).split(".")[0].strip()
    if cp: q["postalcode"] = cp
    intentos = [True, False] if viewbox else [False]
    for bounded in intentos:
        try:
            loc = gcode(q, country_codes="es", addressdetails=True,
                        viewbox=viewbox, bounded=bounded)
        except Exception:
            loc = None
        if loc:
            town = _town(loc)
            match = (not ciudad) or _norm(ciudad) in _norm(town) or _norm(town) in _norm(ciudad)
            if bounded or match:          # bounded => ya está dentro de la ciudad
                return loc.latitude, loc.longitude, True
    return np.nan, np.nan, False

def geocode_df(df, email, progress=None):
    cache = load_cache()
    gcode = _rl(_geoloc(email))
    lats, lons, oks = [], [], []
    tot = len(df)
    for i, (_, row) in enumerate(df.iterrows()):
        if pd.notna(row.get("lat")) and pd.notna(row.get("lon")):
            la, lo, ok = float(row["lat"]), float(row["lon"]), True
        else:
            key = sig(row)
            if key in cache:
                la, lo = cache[key]; ok = True
            else:
                vb = _city_viewbox(gcode, row.get("ciudad"), row.get("pais"))
                la, lo, ok = geocode_one(gcode, row, vb)
                if ok: cache[key] = (la, lo)
        lats.append(la); lons.append(lo); oks.append(bool(ok) and pd.notna(la))
        if progress: progress.progress((i+1)/tot, text=f"Geocodificando {i+1}/{tot}")
    save_cache(cache)
    df = df.copy(); df["lat"], df["lon"], df["ok"] = lats, lons, oks
    return df

def geocode_base(base_info, email):
    """Coordenadas de la base. La base puede estar FUERA de la ciudad (polígono,
    afueras...), así que aquí NO se acota a la ciudad; solo se sesga por país."""
    gcode = _rl(_geoloc(email))
    partes = [_val(base_info.get(k)) for k in ("direccion_base","codigo_postal","ciudad","pais")]
    q = ", ".join([p for p in partes if p])
    if not q:
        return None
    try:
        loc = gcode(q, country_codes="es", addressdetails=True)
    except Exception:
        loc = None
    if loc:
        return (loc.latitude, loc.longitude)
    # segundo intento: estructurado, sin acotar
    row = {"direccion": base_info.get("direccion_base"), "codigo_postal": base_info.get("codigo_postal"),
           "ciudad": base_info.get("ciudad"), "pais": base_info.get("pais")}
    la, lo, ok = geocode_one(gcode, row, None)
    return (la, lo) if ok else None

def geocode_point(query, email):
    cache = load_cache()
    if query in cache: return cache[query]
    gcode = _rl(_geoloc(email))
    try: loc = gcode(query, country_codes="es")
    except Exception: loc = None
    if loc:
        cache[query] = (loc.latitude, loc.longitude); save_cache(cache); return (loc.latitude, loc.longitude)
    return None


# ---------------- lectura de plantillas ----------------
def read_orders(file):
    xls = pd.ExcelFile(file)
    df = pd.read_excel(xls, sheet_name=xls.sheet_names[0])
    for c in COLS:
        if c not in df.columns: df[c] = np.nan
    df = df.dropna(subset=["direccion"])
    return df[COLS]

def read_base(file):
    try:
        b = pd.read_excel(pd.ExcelFile(file), sheet_name="BASE", header=None)
    except Exception:
        return {}
    d = {}
    for _, row in b.iterrows():
        k = str(row[0]).strip().lower()
        v = row[1] if len(row) > 1 else None
        for key in ("direccion_base","codigo_postal","ciudad","pais","hora_inicio","lat_base","lon_base"):
            if k == key and pd.notna(v): d[key] = v
    return d

def hhmm(m): return f"{int(m)//60:02d}:{int(m)%60:02d}"

def to_min(x):
    if pd.isna(x): return None
    try:
        t = pd.to_datetime(x); return t.hour*60 + t.minute
    except Exception: return None


# ---------------- colores / mapas ----------------
def color(i, k):
    import colorsys
    r,g,b = colorsys.hsv_to_rgb(((i/max(k,1))+0.05)%1.0, 0.58, 0.82)
    return f"#{int(r*255):02x}{int(g*255):02x}{int(b*255):02x}"

def hull(pts):
    pts = sorted(set(map(tuple, pts)))
    if len(pts) <= 2: return pts
    def cr(o,a,b): return (a[0]-o[0])*(b[1]-o[1])-(a[1]-o[1])*(b[0]-o[0])
    lo=[]
    for p in pts:
        while len(lo)>=2 and cr(lo[-2],lo[-1],p)<=0: lo.pop()
        lo.append(p)
    up=[]
    for p in reversed(pts):
        while len(up)>=2 and cr(up[-2],up[-1],p)<=0: up.pop()
        up.append(p)
    return lo[:-1]+up[:-1]

def render(m, h=560): components.html(m.get_root().render(), height=h)

def route_km(base, r, df):
    seq = [base] + [(df.lat.iloc[s["idx"]], df.lon.iloc[s["idx"]]) for s in r.stops] + [base]
    return sum(rc.haversine_m(seq[i], seq[i+1]) for i in range(len(seq)-1)) / 1000

def base_marker(m, base):
    folium.Marker(base, tooltip="Base",
        icon=folium.DivIcon(html='<div style="background:#111;color:#fff;border-radius:6px;'
        'width:26px;height:26px;line-height:26px;text-align:center;font:600 13px sans-serif;'
        'box-shadow:0 1px 3px rgba(0,0,0,.5)">B</div>')).add_to(m)

def map_jefe(base, plan, df):
    m = folium.Map(location=base, zoom_start=13, tiles="cartodbpositron")
    base_marker(m, base)
    for r in plan.routes:
        col = color(r.vehicle, plan.k)
        pts = [(df.lat.iloc[s["idx"]], df.lon.iloc[s["idx"]]) for s in r.stops]
        h = hull(pts)
        if len(h) >= 3:
            folium.Polygon([[a,b] for a,b in h], color=col, weight=2, fill=True,
                           fill_color=col, fill_opacity=0.10,
                           tooltip=f"Repartidor {r.vehicle+1} · fin {hhmm(r.end_min)}").add_to(m)
        for s in r.stops:
            tipo = df.tipo.iloc[s["idx"]]
            entrega = str(tipo).startswith("entrega")
            folium.CircleMarker([df.lat.iloc[s["idx"]], df.lon.iloc[s["idx"]]], radius=4,
                color=col, weight=2, fill=True, fill_color=col if entrega else "#ffffff",
                fill_opacity=0.9 if entrega else 1.0,
                tooltip=f"{df.direccion.iloc[s['idx']]} ({tipo})").add_to(m)
    return m

def map_repartidor(base, r, df, geometry, blocks, col):
    seq = [base] + [(df.lat.iloc[s["idx"]], df.lon.iloc[s["idx"]]) for s in r.stops] + [base]
    m = folium.Map(location=base, zoom_start=14, tiles="cartodbpositron")
    if geometry:
        AntPath([[a,b] for a,b in geometry], color=col, weight=5, opacity=0.9,
                delay=700, dash_array=[12,24], pulse_color="#ffffff").add_to(m)
    else:
        folium.PolyLine([[a,b] for a,b in seq], color=col, weight=3, opacity=0.85, dash_array="6,8").add_to(m)
    base_marker(m, base)
    for order, s in enumerate(r.stops, start=1):
        i = s["idx"]; tipo = str(df.tipo.iloc[i]); entrega = tipo.startswith("entrega")
        prio = str(df.prioritario.iloc[i]).lower().startswith("s")
        bg = col if entrega else "#7A4FD0"
        ring = "border:2px solid #D2553F;" if prio else ""
        folium.Marker([df.lat.iloc[i], df.lon.iloc[i]],
            tooltip=f"{order}. {df.direccion.iloc[i]} · {tipo} · llega {hhmm(s['arrival_min'])}",
            icon=folium.DivIcon(html=f'<div style="{ring}background:{bg};color:#fff;border-radius:'
            f'{"50%" if entrega else "4px"};width:22px;height:22px;line-height:22px;text-align:center;'
            f'font:600 11px sans-serif">{order}</div>')).add_to(m)
    for la,lo,rad in (blocks or []):
        folium.Circle([la,lo], radius=rad, color="#D2553F", weight=2, fill=True,
                      fill_color="#D2553F", fill_opacity=0.25, tooltip="Corte").add_to(m)
    return m


# ================= BARRA LATERAL =================
st.sidebar.title("Configuración")

def plantilla_bytes():
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.worksheet.datavalidation import DataValidation
    wb = Workbook(); ws = wb.active; ws.title = "PEDIDOS"
    ws.append(COLS)
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor="1F3864")
    for e in [
        ["P001", "entrega", "Calle Torrente 5", "41010", "Sevilla", "España", "no", "", "", "", "Timbre 2B"],
        ["P002", "recogida", "Calle Feria 45", "41003", "Sevilla", "España", "si", "2026-07-05 12:00", "", "", "Recoger 1 caja"],
    ]:
        ws.append(e)
    dv1 = DataValidation(type="list", formula1='"entrega,recogida"'); ws.add_data_validation(dv1); dv1.add("B2:B1000")
    dv2 = DataValidation(type="list", formula1='"si,no"'); ws.add_data_validation(dv2); dv2.add("G2:G1000")
    for i, w in enumerate([8, 12, 32, 12, 12, 10, 10, 18, 8, 8, 22], 1):
        ws.column_dimensions[chr(64 + i)].width = w
    b = wb.create_sheet("BASE")
    for i, (k, v) in enumerate([("direccion_base", "Calle Industria 10"), ("codigo_postal", "41007"),
                                ("ciudad", "Sevilla"), ("pais", "España"), ("hora_inicio", "08:00"),
                                ("lat_base", ""), ("lon_base", "")], 1):
        b.cell(i, 1, k).font = Font(bold=True); b.cell(i, 2, v)
    b.column_dimensions["A"].width = 20; b.column_dimensions["B"].width = 32
    ins = wb.create_sheet("INSTRUCCIONES")
    ins.column_dimensions["A"].width = 22; ins.column_dimensions["B"].width = 90
    for i, (k, v) in enumerate([
        ("Regla", "Un pedido por fila. No cambies los nombres de las columnas."),
        ("tipo", "entrega o recogida (usa el desplegable)."),
        ("direccion", "Calle y número exactos, por ejemplo Calle Torrente 5."),
        ("codigo_postal", "Muy importante para ubicar bien en el mapa, por ejemplo 41010."),
        ("prioritario", "si o no. Si es si, rellena fecha_hora_limite."),
        ("fecha_hora_limite", "Formato AAAA-MM-DD HH:MM, por ejemplo 2026-07-05 12:00."),
        ("lat / lon", "Opcional. Si ya conoces la coordenada exacta, ponla."),
        ("Hoja BASE", "Rellena la dirección de la base y la hora de salida."),
    ], 1):
        ins.cell(i, 1, k).font = Font(bold=True); ins.cell(i, 2, v).alignment = Alignment(wrap_text=True)
    bio = io.BytesIO(); wb.save(bio); return bio.getvalue()

st.sidebar.download_button("Descargar plantilla de Excel", data=plantilla_bytes(),
                           file_name="plantilla_pedidos.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           use_container_width=True)
st.sidebar.caption("Descarga la plantilla, rellénala y súbela como entregas o recogidas.")

up_e = st.sidebar.file_uploader("Plantilla de ENTREGAS (.xlsx)", type=["xlsx"], key="e")
up_r = st.sidebar.file_uploader("Plantilla de RECOGIDAS (.xlsx)", type=["xlsx"], key="r")
email = "reparto@tesis.local"
if os.path.exists(CACHE):
    usar_cache = st.sidebar.checkbox("Usar direcciones ya geocodificadas", value=False)
else:
    usar_cache = False

st.sidebar.divider()
modo = st.sidebar.radio("Repartidores", ["Fijar cuántos", "Calcular el mínimo"])
if modo == "Fijar cuántos":
    kfix = st.sidebar.slider("Nº de repartidores", 1, 20, 3); kmax = kfix
else:
    kmax = st.sidebar.slider("Máximo de repartidores a probar", 2, 20, 8); kfix = None

with st.sidebar.expander("Jornada y tiempos", expanded=True):
    hora_inicio = st.text_input("Hora de salida (HH:MM)", value="08:00")
    jornada_h = st.slider("Duración de la jornada (h)", 2, 12, 8)
    speed = st.slider("Velocidad media (km/h)", 8, 45, 22)
    service = st.slider("Tiempo por parada (min)", 1, 15, 4)
    tsp_time = st.slider("Segundos de cálculo", 2, 20, 6)

with st.sidebar.expander("Ruta por calles (vista repartidor)"):
    provider = st.radio("Motor de ruta", ["Ninguno", "OpenRouteService"])
    if provider.startswith("Open"):
        ors_key = st.text_input("Clave ORS", type="password",
                                value=_secret("ors_key") or DEFAULT_ORS_KEY)
        st.caption("Viene una clave por defecto. Si deja de funcionar, regístrate gratis en "
                   "https://openrouteservice.org/dev/#/signup y pega aquí tu propia clave.")
    else:
        ors_key = ""
    blocked_text = st.text_area("Cortes de calle (una por línea)", value="", height=70)
    block_radius = st.slider("Radio del corte (m)", 40, 400, 130, step=10)

calcular = st.sidebar.button("Calcular ruta", use_container_width=True)


# ================= CARGA =================
st.title("Planificador de reparto — entregas y recogidas")

def start_minutes(txt):
    try: h,mn = txt.split(":"); return int(h)*60+int(mn)
    except Exception: return 480

if up_e or up_r:
    frames, base_info = [], {}
    if up_e:
        frames.append(read_orders(up_e)); base_info = read_base(up_e) or base_info
    if up_r:
        frames.append(read_orders(up_r)); base_info = base_info or read_base(up_r)
    orders = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame(columns=COLS)
    st.session_state.orders = orders
    st.session_state.base_info = base_info
    st.caption(f"Cargados {len(orders)} pedidos "
               f"({(orders.tipo.astype(str).str.startswith('entrega')).sum()} entregas, "
               f"{(orders.tipo.astype(str).str.startswith('recogida')).sum()} recogidas).")
else:
    st.info("Sube la plantilla de entregas y/o la de recogidas en la barra lateral.")

# ================= CÁLCULO =================
if calcular and st.session_state.get("orders") is not None:
    orders = st.session_state.orders.copy()
    base_info = st.session_state.get("base_info", {})
    prog = st.progress(0.0, text="Geocodificando…")
    if usar_cache:
        cache = load_cache()
        orders["lat"] = orders.apply(lambda r: cache.get(sig(r),(np.nan,np.nan))[0], axis=1)
        orders["lon"] = orders.apply(lambda r: cache.get(sig(r),(np.nan,np.nan))[1], axis=1)
        orders["ok"] = orders.lat.notna()
    else:
        orders = geocode_df(orders, email, prog)
    prog.empty()

    # base (acotada a la ciudad; si no se encuentra, se avisa y no se coloca)
    if base_info.get("lat_base") and base_info.get("lon_base"):
        base = (float(base_info["lat_base"]), float(base_info["lon_base"]))
    else:
        base = geocode_base(base_info, email)
    if base is None:
        st.error("No pude ubicar la BASE dentro de la ciudad indicada. Revisa 'direccion_base' "
                 "(calle y número) o pon 'lat_base' y 'lon_base' en la hoja BASE.")
        st.dataframe(pd.DataFrame([{
            "direccion_base": base_info.get("direccion_base"),
            "codigo_postal": base_info.get("codigo_postal"),
            "ciudad": base_info.get("ciudad"), "pais": base_info.get("pais")}]),
            use_container_width=True, hide_index=True)
        st.stop()

    st.session_state.unresolved = orders.loc[~orders.ok, COLS].reset_index(drop=True)
    ok = orders[orders.ok].reset_index(drop=True)
    if len(ok) < 1:
        st.error("No hay pedidos geocodificados."); st.stop()

    start_min = start_minutes(hora_inicio)
    stops = []
    for _, r in ok.iterrows():
        dl = to_min(r["fecha_hora_limite"]) if str(r["prioritario"]).lower().startswith("s") else None
        stops.append({"lat":float(r.lat),"lon":float(r.lon),"tipo":str(r.tipo),"deadline_min":dl})

    kw = dict(start_min=start_min, speed_kmh=speed, service_min=service,
              shift_min=jornada_h*60, time_limit_s=tsp_time)
    with st.spinner("Optimizando rutas (equilibrio por tiempo)…"):
        plan = rc.find_min_k(stops, base, kmax, **kw) if kfix is None else rc.solve_vrp(stops, base, kfix, **kw)
    st.session_state.plan = plan
    st.session_state.pdf = ok
    st.session_state.base = base

# ================= VISTAS =================
plan = st.session_state.get("plan"); pdf = st.session_state.get("pdf"); base = st.session_state.get("base")
if plan is not None and pdf is not None:
    unres = st.session_state.get("unresolved")
    if unres is not None and len(unres):
        st.warning(f"{len(unres)} dirección(es) no se pudieron ubicar dentro de la ciudad indicada. "
                   "Revisa el número y el código postal.")
        with st.expander("Ver direcciones no ubicadas"):
            st.dataframe(unres[["id","tipo","direccion","codigo_postal","ciudad"]],
                         use_container_width=True, hide_index=True)
    if plan.dropped:
        st.warning(f"{len(plan.dropped)} pedido(s) no caben en la jornada/plazos con esta flota. "
                   "Sube el nº de repartidores o la jornada.")
        with st.expander("Ver detalle de los pedidos no asignados"):
            det = pdf.iloc[plan.dropped][["id","tipo","direccion","codigo_postal",
                                          "prioritario","fecha_hora_limite"]].reset_index(drop=True)
            st.dataframe(det, use_container_width=True, hide_index=True)
    tab1, tab2 = st.tabs(["Vista del jefe", "Vista del repartidor"])

    with tab1:
        mins = [r.route_min for r in plan.routes] or [0]
        # prioritarios cumplidos
        pri_ok = pri_tot = 0
        for r in plan.routes:
            for s in r.stops:
                row = pdf.iloc[s["idx"]]
                if str(row.prioritario).lower().startswith("s") and pd.notna(row.fecha_hora_limite):
                    pri_tot += 1
                    if s["arrival_min"] <= (to_min(row.fecha_hora_limite) or 1e9): pri_ok += 1
        km_total = sum(route_km(base, r, pdf) for r in plan.routes)
        c = st.columns(5)
        c[0].metric("Repartidores", plan.k)
        c[1].metric("Equilibrio (dif. máx-mín)", f"{round(max(mins)-min(mins))} min")
        c[2].metric("Jornada más larga", hhmm(plan.makespan_min))
        c[3].metric("Distancia total", f"{km_total:.1f} km")
        c[4].metric("Prioritarios a tiempo", f"{pri_ok}/{pri_tot}")
        render(map_jefe(base, plan, pdf))
        resumen = pd.DataFrame([{
            "Repartidor": f"R{r.vehicle+1}", "Entregas": r.n_entregas, "Recogidas": r.n_recogidas,
            "Distancia": f"{route_km(base, r, pdf):.1f} km",
            "Duración": f"{round(r.route_min)} min", "Hora de fin": hhmm(r.end_min),
        } for r in plan.routes])
        st.dataframe(resumen, use_container_width=True, hide_index=True)

    with tab2:
        opt = {f"Repartidor {r.vehicle+1}  ·  {len(r.stops)} paradas  ·  fin {hhmm(r.end_min)}": r for r in plan.routes}
        r = opt[st.selectbox("Elige el repartidor", list(opt.keys()))]
        col = color(r.vehicle, plan.k)

        blocks = []
        if provider != "Ninguno" and ors_key and blocked_text.strip():
            for line in [x.strip() for x in blocked_text.splitlines() if x.strip()]:
                pt = geocode_point(f"{line}, Sevilla, España", email)
                if pt: blocks.append((pt[0], pt[1], block_radius))

        geometry = None; road = None
        if provider != "Ninguno" and ors_key:
            seq = [base] + [(pdf.lat.iloc[s["idx"]], pdf.lon.iloc[s["idx"]]) for s in r.stops] + [base]
            with st.spinner("Trazando por calles…"):
                rings = [ors.circle_polygon(a,b,rad) for a,b,rad in blocks]
                road = ors.route_geometry(seq, ors_key, avoid=ors.to_multipolygon(rings))
            if road and road.get("error"): st.warning(road["error"])
            elif road: geometry = road["geometry"]

        if road and road.get("distance_m"):
            km = road["distance_m"] / 1000; km_lbl = "Distancia (carretera)"
        else:
            km = route_km(base, r, pdf); km_lbl = "Distancia"
        a = st.columns(4)
        a[0].metric("Paradas", len(r.stops))
        a[1].metric("Entregas / Recogidas", f"{r.n_entregas} / {r.n_recogidas}")
        a[2].metric(km_lbl, f"{km:.1f} km")
        a[3].metric("Hora de fin", hhmm(r.end_min))
        render(map_repartidor(base, r, pdf, geometry, blocks, col))

        hoja = []
        hoja.append({"Orden":"Salida","Hora":hhmm(plan.start_min),"Tipo":"base","Dirección":"BASE"})
        for order, s in enumerate(r.stops, start=1):
            row = pdf.iloc[s["idx"]]
            pri = " (prioritario)" if str(row.prioritario).lower().startswith("s") else ""
            hoja.append({"Orden":order,"Hora":hhmm(s["arrival_min"]),
                         "Tipo":str(row.tipo)+pri,"Dirección":row.direccion})
        hoja.append({"Orden":"Regreso","Hora":hhmm(r.end_min),"Tipo":"base (descargar recogidas)","Dirección":"BASE"})
        hoja = pd.DataFrame(hoja)
        st.subheader("Hoja de ruta")
        st.dataframe(hoja, use_container_width=True, hide_index=True)
        st.download_button("Descargar hoja de ruta (CSV)", hoja.to_csv(index=False).encode("utf-8"),
                           file_name=f"ruta_R{r.vehicle+1}.csv", mime="text/csv")
elif st.session_state.get("orders") is not None:
    st.info("Pulsa **Calcular ruta** en la barra lateral.")
